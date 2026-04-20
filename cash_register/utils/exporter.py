"""
exporter.py — Excel export logic, fully isolated from the UI.
Uses openpyxl. Call export_to_excel(state, path) from anywhere.
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from cash_register.core.models import LedgerState
from cash_register.utils.formatters import money, money_or_dash

# ── Excel-specific colour stops (hex, no #) ───────────────────────────────────
XL_HEADER_BG  = "a9d6bb"   # gum leaf
XL_HEADER_FG  = "1e3a22"   # dark forest text
XL_OPENING_BG = "a7e7b4"   # chinook
XL_FOOTER_BG  = "d2e9af"   # caper
XL_ODD_BG     = "ffffff"
XL_EVEN_BG    = "e3f2d4"   # chrome white
XL_BORDER_CLR = "c5ddb5"


def _border() -> Border:
    side = Side(style="thin", color=XL_BORDER_CLR)
    return Border(left=side, right=side, top=side, bottom=side)


def export_to_excel(state: LedgerState, path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Cash Register"

    bdr       = _border()
    c_align   = Alignment(horizontal="center", vertical="center")
    l_align   = Alignment(horizontal="left",   vertical="center")
    r_align   = Alignment(horizontal="right",  vertical="center")

    def font(bold=False, italic=False, size=10, color=XL_HEADER_FG):
        return Font(name="Segoe UI", bold=bold, italic=italic,
                    size=size, color=color)

    def fill(hex_color: str) -> PatternFill:
        return PatternFill("solid", fgColor=hex_color)

    def apply(cell, value, *, fnt, fll, aln, row_h=None):
        cell.value     = value
        cell.font      = fnt
        cell.fill      = fll
        cell.alignment = aln
        cell.border    = bdr

    # ── title row ─────────────────────────────────────────────────────────────
    ws.merge_cells("A1:E1")
    ws["A1"].value     = f"Cash Register — {state.current_date or ''}"
    ws["A1"].font      = font(bold=True, size=13)
    ws["A1"].alignment = c_align
    ws["A1"].fill      = fill(XL_HEADER_BG)
    ws.row_dimensions[1].height = 28

    # ── column headers ────────────────────────────────────────────────────────
    headers = ["Date", "Name", "Credit (CR)", "Debit (DR)", "Cash in Hand"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font      = font(bold=True, color="1e3a22")
        cell.fill      = fill(XL_HEADER_BG)
        cell.alignment = c_align
        cell.border    = bdr
    ws.row_dimensions[2].height = 22

    # ── opening row ───────────────────────────────────────────────────────────
    opening = state.opening_cash or 0.0
    cur_date = state.current_date or ""
    opening_vals = [cur_date, "Opening Balance",
                    money(opening), "—", money(opening)]
    for col, val in enumerate(opening_vals, 1):
        cell = ws.cell(row=3, column=col, value=val)
        cell.font      = font(italic=True)
        cell.fill      = fill(XL_OPENING_BG)
        cell.border    = bdr
        cell.alignment = (c_align if col == 1 else
                          l_align if col == 2 else r_align)
    ws.row_dimensions[3].height = 20

    # ── transaction rows ──────────────────────────────────────────────────────
    running = opening
    for i, tx in enumerate(state.rows):
        r   = 4 + i
        running += tx.cr - tx.dr
        row_fill = fill(XL_ODD_BG if i % 2 == 0 else XL_EVEN_BG)
        vals = [
            tx.date or cur_date,
            tx.name,
            money_or_dash(tx.cr),
            money_or_dash(tx.dr),
            money(running),
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font      = font()
            cell.fill      = row_fill
            cell.border    = bdr
            cell.alignment = (c_align if col == 1 else
                              l_align if col == 2 else r_align)
        ws.row_dimensions[r].height = 20

    # ── footer row ────────────────────────────────────────────────────────────
    fr = 4 + len(state.rows)
    footer_vals = ["", "Cash in Hand",
                   money(state.total_cr), money(state.total_dr),
                   money(state.cash_in_hand)]
    for col, val in enumerate(footer_vals, 1):
        cell = ws.cell(row=fr, column=col, value=val)
        cell.font      = font(bold=True)
        cell.fill      = fill(XL_FOOTER_BG)
        cell.border    = bdr
        cell.alignment = l_align if col <= 2 else r_align
    ws.row_dimensions[fr].height = 22

    # ── column widths & freeze ────────────────────────────────────────────────
    col_widths = {"A": 14, "B": 30, "C": 18, "D": 18, "E": 20}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A3"

    wb.save(path)
