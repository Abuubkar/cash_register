import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog
import json
import os
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DATA_FILE = os.path.join(os.path.expanduser("~"), "cash_register_data.json")

# ── colours ──────────────────────────────────────────────────────────────────
BG          = "#F5F5F0"
HEADER_BG   = "#2C2C2A"
HEADER_FG   = "#FFFFFF"
ROW_ODD     = "#FFFFFF"
ROW_EVEN    = "#F8F8F4"
OPENING_BG  = "#E1F5EE"
FOOTER_BG   = "#FAEEDA"
ACCENT      = "#1D9E75"
ACCENT_DARK = "#0F6E56"
BTN_RED     = "#A32D2D"
BTN_RED_D   = "#791F1F"
BORDER      = "#D3D1C7"
TEXT_PRI    = "#2C2C2A"
TEXT_SEC    = "#5F5E5A"

def load_data():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, "r") as f:
            return json.load(f)
    return {"opening_cash": None, "current_date": None, "rows": []}

def save_data(data):
    with open(DATA_FILE, "w") as f:
        json.dump(data, f, indent=2)


class CashRegisterApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Cash Register")
        self.geometry("900x620")
        self.minsize(750, 500)
        self.configure(bg=BG)
        self.resizable(True, True)

        self.data = load_data()
        self._setup_styles()
        self._build_ui()

        # first run: ask for opening cash
        if self.data["opening_cash"] is None:
            self.after(200, self._ask_opening_cash)
        else:
            self._refresh_table()

    # ── styles ────────────────────────────────────────────────────────────────
    def _setup_styles(self):
        style = ttk.Style(self)
        style.theme_use("clam")

        style.configure("Treeview",
            background=ROW_ODD, foreground=TEXT_PRI,
            fieldbackground=ROW_ODD, rowheight=34,
            font=("Segoe UI", 10), borderwidth=0)
        style.configure("Treeview.Heading",
            background=HEADER_BG, foreground=HEADER_FG,
            font=("Segoe UI", 10, "bold"), relief="flat", padding=(8, 6))
        style.map("Treeview",
            background=[("selected", ACCENT)],
            foreground=[("selected", "#FFFFFF")])
        style.map("Treeview.Heading", background=[("active", "#444441")])

        style.configure("Accent.TButton",
            background=ACCENT, foreground="#FFFFFF",
            font=("Segoe UI", 10, "bold"), relief="flat",
            padding=(14, 8), borderwidth=0)
        style.map("Accent.TButton",
            background=[("active", ACCENT_DARK), ("pressed", ACCENT_DARK)])

        style.configure("Danger.TButton",
            background=BTN_RED, foreground="#FFFFFF",
            font=("Segoe UI", 10), relief="flat",
            padding=(14, 8), borderwidth=0)
        style.map("Danger.TButton",
            background=[("active", BTN_RED_D), ("pressed", BTN_RED_D)])

        style.configure("Flat.TButton",
            background="#E0DDD4", foreground=TEXT_PRI,
            font=("Segoe UI", 10), relief="flat",
            padding=(14, 8), borderwidth=0)
        style.map("Flat.TButton",
            background=[("active", BORDER)])

    # ── UI ────────────────────────────────────────────────────────────────────
    def _build_ui(self):
        # title bar
        title_frame = tk.Frame(self, bg=HEADER_BG, height=56)
        title_frame.pack(fill="x")
        title_frame.pack_propagate(False)
        tk.Label(title_frame, text="💼  Cash Register",
                 bg=HEADER_BG, fg=HEADER_FG,
                 font=("Segoe UI", 14, "bold")).pack(side="left", padx=20, pady=14)

        # toolbar
        toolbar = tk.Frame(self, bg=BG, pady=10, padx=16)
        toolbar.pack(fill="x")

        # date selector
        tk.Label(toolbar, text="Date:", bg=BG, fg=TEXT_SEC,
                 font=("Segoe UI", 10)).pack(side="left")
        self.date_var = tk.StringVar(value=self.data.get("current_date") or str(date.today()))
        self.date_entry = tk.Entry(toolbar, textvariable=self.date_var,
                                   font=("Segoe UI", 10), width=13,
                                   bg="white", fg=TEXT_PRI, relief="solid",
                                   bd=1, highlightthickness=0)
        self.date_entry.pack(side="left", padx=(4, 12), ipady=4)

        ttk.Button(toolbar, text="Set Date", style="Accent.TButton",
                   command=self._set_date).pack(side="left", padx=(0, 20))

        ttk.Button(toolbar, text="＋ Add Row", style="Flat.TButton",
                   command=self._open_add_row).pack(side="left", padx=(0, 6))
        ttk.Button(toolbar, text="✎ Edit Row", style="Flat.TButton",
                   command=self._open_edit_row).pack(side="left", padx=(0, 6))
        ttk.Button(toolbar, text="✕ Delete Row", style="Danger.TButton",
                   command=self._delete_row).pack(side="left", padx=(0, 20))
        ttk.Button(toolbar, text="⬇ Export Excel", style="Accent.TButton",
                   command=self._export_excel).pack(side="right")

        # divider
        tk.Frame(self, bg=BORDER, height=1).pack(fill="x")

        # table
        table_frame = tk.Frame(self, bg=BG)
        table_frame.pack(fill="both", expand=True, padx=16, pady=12)

        cols = ("date", "name", "cr", "dr", "cash_in_hand")
        self.tree = ttk.Treeview(table_frame, columns=cols, show="headings",
                                  selectmode="browse")

        headers = {"date": "Date", "name": "Name",
                   "cr": "Credit (CR)", "dr": "Debit (DR)", "cash_in_hand": "Cash in Hand"}
        widths   = {"date": 110, "name": 220, "cr": 130, "dr": 130, "cash_in_hand": 140}
        anchors  = {"date": "center", "name": "w", "cr": "e", "dr": "e", "cash_in_hand": "e"}

        for c in cols:
            self.tree.heading(c, text=headers[c])
            self.tree.column(c, width=widths[c], anchor=anchors[c], minwidth=80)

        # tag styles
        self.tree.tag_configure("opening", background=OPENING_BG, font=("Segoe UI", 10, "italic"))
        self.tree.tag_configure("footer",  background=FOOTER_BG,  font=("Segoe UI", 10, "bold"))
        self.tree.tag_configure("odd",     background=ROW_ODD)
        self.tree.tag_configure("even",    background=ROW_EVEN)

        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        self.tree.pack(fill="both", expand=True)

        # status bar
        status_bar = tk.Frame(self, bg="#E8E6DF", height=28)
        status_bar.pack(fill="x", side="bottom")
        status_bar.pack_propagate(False)
        self.status_var = tk.StringVar(value="Ready")
        tk.Label(status_bar, textvariable=self.status_var,
                 bg="#E8E6DF", fg=TEXT_SEC,
                 font=("Segoe UI", 9), anchor="w").pack(side="left", padx=12, pady=4)

    # ── first run ─────────────────────────────────────────────────────────────
    def _ask_opening_cash(self):
        dlg = EntryDialog(self, title="Opening Balance",
                          prompt="Enter your current Cash in Hand to get started:",
                          label="Opening Cash in Hand (Rs)")
        val = dlg.result
        if val is None:
            self.data["opening_cash"] = 0.0
        else:
            self.data["opening_cash"] = val
        self.data["current_date"] = self.date_var.get()
        self.data["rows"] = []
        save_data(self.data)
        self._refresh_table()
        self._set_status(f"Opening cash set to Rs {self.data['opening_cash']:,.2f}")

    # ── date change ───────────────────────────────────────────────────────────
    def _set_date(self):
        new_date = self.date_var.get().strip()
        if not new_date:
            messagebox.showwarning("Invalid", "Please enter a date.")
            return
        old_date = self.data.get("current_date")
        if new_date == old_date:
            messagebox.showinfo("Same Date", "That is already the current date.")
            return
        # confirm
        last_cash = self._calc_cash_in_hand()
        confirm = messagebox.askyesno(
            "Change Date",
            f"Changing date to {new_date}.\n\n"
            f"All rows for '{old_date}' will be cleared.\n"
            f"Closing cash in hand (Rs {last_cash:,.2f}) will carry forward as opening balance.\n\n"
            "Continue?")
        if not confirm:
            return

        self.data["opening_cash"] = last_cash
        self.data["current_date"] = new_date
        self.data["rows"] = []
        save_data(self.data)
        self._refresh_table()
        self._set_status(f"Date changed to {new_date}. Opening balance: Rs {last_cash:,.2f}")

    # ── add / edit row dialogs ─────────────────────────────────────────────────
    def _open_add_row(self):
        dlg = RowDialog(self, title="Add Transaction",
                        date_val=self.data.get("current_date", ""),
                        name_val="", cr_val="", dr_val="")
        if dlg.result:
            self.data["rows"].append(dlg.result)
            save_data(self.data)
            self._refresh_table()
            self._set_status("Row added.")

    def _open_edit_row(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Select Row", "Please select a transaction row to edit.")
            return
        item = self.tree.item(sel[0])
        tags = item["tags"]
        if "opening" in tags or "footer" in tags:
            messagebox.showinfo("Cannot Edit", "The opening balance and footer rows cannot be edited.")
            return
        idx = int(self.tree.item(sel[0], "values")[5])  # hidden index
        row = self.data["rows"][idx]
        dlg = RowDialog(self, title="Edit Transaction",
                        date_val=row["date"], name_val=row["name"],
                        cr_val=row["cr"], dr_val=row["dr"])
        if dlg.result:
            self.data["rows"][idx] = dlg.result
            save_data(self.data)
            self._refresh_table()
            self._set_status("Row updated.")

    def _delete_row(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Select Row", "Please select a transaction row to delete.")
            return
        item = self.tree.item(sel[0])
        tags = item["tags"]
        if "opening" in tags or "footer" in tags:
            messagebox.showinfo("Cannot Delete", "The opening balance and footer rows cannot be deleted.")
            return
        idx = int(self.tree.item(sel[0], "values")[5])
        confirm = messagebox.askyesno("Delete Row", "Delete this transaction?")
        if confirm:
            self.data["rows"].pop(idx)
            save_data(self.data)
            self._refresh_table()
            self._set_status("Row deleted.")

    # ── table refresh ─────────────────────────────────────────────────────────
    def _calc_cash_in_hand(self):
        opening = self.data.get("opening_cash") or 0.0
        total_cr = sum(r.get("cr", 0) or 0 for r in self.data["rows"])
        total_dr = sum(r.get("dr", 0) or 0 for r in self.data["rows"])
        return opening + total_cr - total_dr

    def _refresh_table(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

        opening = self.data.get("opening_cash") or 0.0
        cur_date = self.data.get("current_date", "")

        # cols: date, name, cr, dr, cash_in_hand, (hidden_idx)
        # Row 1 – opening balance
        self.tree["columns"] = ("date", "name", "cr", "dr", "cash_in_hand", "idx")
        self.tree.column("idx", width=0, minwidth=0, stretch=False)
        self.tree.heading("idx", text="")

        self.tree.insert("", "end",
            values=(cur_date, "Opening Balance",
                    f"Rs {opening:,.2f}", "—",
                    f"Rs {opening:,.2f}", -1),
            tags=("opening",))

        running = opening
        for i, row in enumerate(self.data["rows"]):
            cr = row.get("cr") or 0.0
            dr = row.get("dr") or 0.0
            running = running + cr - dr
            tag = "odd" if i % 2 == 0 else "even"
            cr_str = f"Rs {cr:,.2f}" if cr else "—"
            dr_str = f"Rs {dr:,.2f}" if dr else "—"
            self.tree.insert("", "end",
                values=(row.get("date", cur_date),
                        row.get("name", ""),
                        cr_str, dr_str,
                        f"Rs {running:,.2f}", i),
                tags=(tag,))

        # footer
        total_cr = sum(r.get("cr", 0) or 0 for r in self.data["rows"])
        total_dr = sum(r.get("dr", 0) or 0 for r in self.data["rows"])
        cash = opening + total_cr - total_dr
        self.tree.insert("", "end",
            values=("", "Cash in Hand",
                    f"Rs {total_cr:,.2f}", f"Rs {total_dr:,.2f}",
                    f"Rs {cash:,.2f}", -2),
            tags=("footer",))

        self._set_status(
            f"Date: {cur_date}  |  "
            f"Opening: Rs {opening:,.2f}  |  "
            f"Cash in Hand: Rs {cash:,.2f}  |  "
            f"Transactions: {len(self.data['rows'])}")

    # ── export ────────────────────────────────────────────────────────────────
    def _export_excel(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"cash_register_{self.data.get('current_date','')}.xlsx",
            title="Export to Excel")
        if not path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Cash Register"

        # styles
        thin = Side(style="thin", color="D3D1C7")
        bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

        hdr_font  = Font(name="Segoe UI", bold=True, color="FFFFFF", size=11)
        hdr_fill  = PatternFill("solid", fgColor="2C2C2A")
        hdr_align = Alignment(horizontal="center", vertical="center")

        open_fill  = PatternFill("solid", fgColor="E1F5EE")
        footer_fill= PatternFill("solid", fgColor="FAEEDA")
        odd_fill   = PatternFill("solid", fgColor="FFFFFF")
        even_fill  = PatternFill("solid", fgColor="F8F8F4")

        bold_font = Font(name="Segoe UI", bold=True, size=10)
        norm_font = Font(name="Segoe UI", size=10)
        ital_font = Font(name="Segoe UI", italic=True, size=10)

        r_align = Alignment(horizontal="right", vertical="center")
        l_align = Alignment(horizontal="left",  vertical="center")
        c_align = Alignment(horizontal="center", vertical="center")

        # title row
        ws.merge_cells("A1:E1")
        ws["A1"] = f"Cash Register — {self.data.get('current_date', '')}"
        ws["A1"].font = Font(name="Segoe UI", bold=True, size=13, color="2C2C2A")
        ws["A1"].alignment = c_align
        ws.row_dimensions[1].height = 28

        # header
        headers = ["Date", "Name", "Credit (CR)", "Debit (DR)", "Cash in Hand"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font     = hdr_font
            cell.fill     = hdr_fill
            cell.alignment= hdr_align
            cell.border   = bdr
        ws.row_dimensions[2].height = 22

        opening = self.data.get("opening_cash") or 0.0
        cur_date = self.data.get("current_date", "")

        def money(v):
            return f"Rs {v:,.2f}"

        # opening row
        for col, val in enumerate([cur_date, "Opening Balance", money(opening), "—", money(opening)], 1):
            cell = ws.cell(row=3, column=col, value=val)
            cell.font      = ital_font
            cell.fill      = open_fill
            cell.border    = bdr
            cell.alignment = r_align if col >= 3 else (c_align if col == 1 else l_align)
        ws.row_dimensions[3].height = 20

        running = opening
        data_rows = self.data["rows"]
        for i, row in enumerate(data_rows):
            r = 4 + i
            cr = row.get("cr") or 0.0
            dr = row.get("dr") or 0.0
            running = running + cr - dr
            fill = odd_fill if i % 2 == 0 else even_fill
            vals = [row.get("date", cur_date), row.get("name", ""),
                    money(cr) if cr else "—", money(dr) if dr else "—", money(running)]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=r, column=col, value=val)
                cell.font      = norm_font
                cell.fill      = fill
                cell.border    = bdr
                cell.alignment = r_align if col >= 3 else (c_align if col == 1 else l_align)
            ws.row_dimensions[r].height = 20

        # footer
        fr = 4 + len(data_rows)
        total_cr = sum(r.get("cr", 0) or 0 for r in data_rows)
        total_dr = sum(r.get("dr", 0) or 0 for r in data_rows)
        cash     = opening + total_cr - total_dr
        footer_vals = ["", "Cash in Hand", money(total_cr), money(total_dr), money(cash)]
        for col, val in enumerate(footer_vals, 1):
            cell = ws.cell(row=fr, column=col, value=val)
            cell.font      = bold_font
            cell.fill      = footer_fill
            cell.border    = bdr
            cell.alignment = r_align if col >= 3 else l_align
        ws.row_dimensions[fr].height = 22

        # column widths
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 28
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 20

        # freeze header rows
        ws.freeze_panes = "A3"

        wb.save(path)
        messagebox.showinfo("Exported", f"Excel file saved to:\n{path}")
        self._set_status(f"Exported to {path}")

    def _set_status(self, msg):
        self.status_var.set(msg)


# ── dialogs ───────────────────────────────────────────────────────────────────
class EntryDialog(tk.Toplevel):
    def __init__(self, parent, title, prompt, label):
        super().__init__(parent)
        self.title(title)
        self.result = None
        self.resizable(False, False)
        self.grab_set()
        self.configure(bg=BG)
        self._build(prompt, label)
        self.transient(parent)
        self.wait_window()

    def _build(self, prompt, label):
        tk.Label(self, text=prompt, bg=BG, fg=TEXT_PRI,
                 font=("Segoe UI", 10), wraplength=320,
                 justify="left").pack(padx=24, pady=(20, 8), anchor="w")
        tk.Label(self, text=label, bg=BG, fg=TEXT_SEC,
                 font=("Segoe UI", 9)).pack(padx=24, anchor="w")
        self.entry = tk.Entry(self, font=("Segoe UI", 11), width=22,
                              bg="white", fg=TEXT_PRI,
                              relief="solid", bd=1)
        self.entry.pack(padx=24, pady=(4, 16), ipady=5)
        self.entry.focus_set()
        self.entry.bind("<Return>", lambda e: self._ok())

        btn_frame = tk.Frame(self, bg=BG)
        btn_frame.pack(fill="x", padx=24, pady=(0, 16))
        ttk.Button(btn_frame, text="OK", style="Accent.TButton",
                   command=self._ok).pack(side="right", padx=(6, 0))
        ttk.Button(btn_frame, text="Cancel", style="Flat.TButton",
                   command=self.destroy).pack(side="right")

    def _ok(self):
        try:
            self.result = float(self.entry.get().replace(",", ""))
            self.destroy()
        except ValueError:
            messagebox.showerror("Invalid", "Please enter a valid number.", parent=self)


class RowDialog(tk.Toplevel):
    def __init__(self, parent, title, date_val, name_val, cr_val, dr_val):
        super().__init__(parent)
        self.title(title)
        self.result = None
        self.resizable(False, False)
        self.grab_set()
        self.configure(bg=BG)
        self._build(date_val, name_val, cr_val, dr_val)
        self.transient(parent)
        self.wait_window()

    def _build(self, date_val, name_val, cr_val, dr_val):
        pad = {"padx": 24, "pady": 4}

        def field(label, val, row):
            tk.Label(self, text=label, bg=BG, fg=TEXT_SEC,
                     font=("Segoe UI", 9)).grid(row=row*2, column=0, sticky="w", **pad)
            e = tk.Entry(self, font=("Segoe UI", 10), width=28,
                         bg="white", fg=TEXT_PRI, relief="solid", bd=1)
            e.insert(0, str(val) if val is not None else "")
            e.grid(row=row*2+1, column=0, **pad, ipady=4)
            return e

        self.grid_columnconfigure(0, weight=1)
        self.e_date = field("Date", date_val, 0)
        self.e_name = field("Name / Description", name_val, 1)
        self.e_cr   = field("Credit (CR)  — leave blank if none", cr_val, 2)
        self.e_dr   = field("Debit (DR)   — leave blank if none", dr_val, 3)

        btn_frame = tk.Frame(self, bg=BG)
        btn_frame.grid(row=9, column=0, padx=24, pady=(12, 16), sticky="e")
        ttk.Button(btn_frame, text="Save", style="Accent.TButton",
                   command=self._save).pack(side="right", padx=(6, 0))
        ttk.Button(btn_frame, text="Cancel", style="Flat.TButton",
                   command=self.destroy).pack(side="right")

        self.e_name.focus_set()
        self.bind("<Return>", lambda e: self._save())

    def _save(self):
        date_v = self.e_date.get().strip()
        name_v = self.e_name.get().strip()
        cr_s   = self.e_cr.get().strip().replace(",", "")
        dr_s   = self.e_dr.get().strip().replace(",", "")

        if not name_v:
            messagebox.showerror("Required", "Name/Description is required.", parent=self)
            return

        cr = None
        dr = None
        try:
            if cr_s:
                cr = float(cr_s)
                if cr < 0:
                    raise ValueError
        except ValueError:
            messagebox.showerror("Invalid", "Credit must be a positive number.", parent=self)
            return
        try:
            if dr_s:
                dr = float(dr_s)
                if dr < 0:
                    raise ValueError
        except ValueError:
            messagebox.showerror("Invalid", "Debit must be a positive number.", parent=self)
            return

        self.result = {"date": date_v, "name": name_v,
                       "cr": cr or 0.0, "dr": dr or 0.0}
        self.destroy()


if __name__ == "__main__":
    app = CashRegisterApp()
    app.mainloop()
